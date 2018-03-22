import os
import json
import time
import requests
import datetime
from openpyxl import load_workbook


import logging
logger = logging.getLogger()

CHAT_ENDPOINT = os.environ.get("CHAT_ENDPOINT", "https://chat.unbabel.com")
CHAT_CUSTOMER = os.environ.get("CHAT_CUSTOMER", "5aa5b1feaa8af31d75698f1e")
CHAT_SLEEP_TIME = float(os.environ.get("CHAT_SLEEP_TIME", 13.5))
CHAT_START_TAB = int(os.environ.get("CHAT_START_TAB", 1))
CHAT_END_TAB = int(os.environ.get("CHAT_END_TAB", 361))


def create_chat(customer_id, chat_name):
    print("Creating chat for {} {}".format(customer_id, chat_name))
    response = requests.post('{}/customers/{}/conversations'.format(
        CHAT_ENDPOINT, customer_id))

    if response.status_code == 201:
        print("Chat successfully created.")
        data_response = response.json()

        response = requests.put(
            '{}/customers/{}/conversations/{}'
            ''.format(CHAT_ENDPOINT, customer_id, data_response['id']),
            data=json.dumps({"name": chat_name}),
            headers={"Content-Type": "application/json"}
        )

        if response.status_code == 200:
            print("Chat name changed to {}".format(chat_name))
            return data_response

        else:
            print("Failed to change name")
            return data_response

    else:
        print("STATUS_CODE {}. Failed because {}"
                     "".format(response.status_code, response.text))
        return None


def create_message(conversation_id, text, source_language, target_language):
    print("Sending message with text: {}".format(text.encode('utf-8')))
    data = {
        "text": text,
        "source_language": source_language,
        "target_language": target_language
    }
    headers = {
        "Content-Type": "application/json"
    }
    response = requests.post(
        '{}/conversations/{}'.format(CHAT_ENDPOINT, conversation_id),
        data=json.dumps(data), headers=headers)

    if response.status_code == 200:
        print("Message successfully created")
        return response.json()
    else:
        print("There was an error in creating the message")
        return None


def run_test():
    print("Running CHAT TEST on {} for customer \"{}\" with a timeout "
                "of \"{}\" seconds".format(CHAT_ENDPOINT, CHAT_CUSTOMER,
                                           CHAT_SLEEP_TIME))
    # generate test tab names     core/management/scripts/
    filepath = os.path.join(os.getcwd(), '20180309_all_agent.xlsx')  # noqa
    easyjet_sheets = load_workbook(filepath)

    sheet_names = ["Test{}".format(sheet_number)
                   for sheet_number in range(CHAT_START_TAB, CHAT_END_TAB)]

    for sheet_name in sheet_names:
        tab_sheet = easyjet_sheets[sheet_name]
        print("------##Test {}##-----".format(sheet_name))

        chat = create_chat(CHAT_CUSTOMER, sheet_name)

        if not chat:
            print("An error occurred while creating a chat for "
                         "customer {}".format(CHAT_CUSTOMER))
            import sys
            sys.exit()

        if 'id' not in chat:
            print("Couldn't retrieve id from chat for customer {}"
                         "".format(CHAT_CUSTOMER))
            import sys
            sys.exit()

        for row in tab_sheet.iter_rows('A{}:A{}'.format(tab_sheet.min_row,
                                                        tab_sheet.max_row)):
            for cell in row:
                if cell.value and cell.value != 'Original text':
                    _ = create_message(chat['id'], cell.value, 'en', 'fr')
                    print("Sleeping {} seconds... {}"
                                "".format(CHAT_SLEEP_TIME,
                                          datetime.datetime.now()))
                    time.sleep(CHAT_SLEEP_TIME)

if __name__ == "__main__":
    run_test()
